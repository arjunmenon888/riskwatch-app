import os
import datetime
import base64
import io

import dash
from dash import dcc, html, Input, Output, State, no_update
from dash.exceptions import PreventUpdate

# Import shared custom modules
import database
from ai_module import get_ai_analysis

# Imports for Excel Generation
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --- Layout Functions for Safety Observation ---

def build_observation_form_page():
    """Builds the layout for the observation submission form."""
    return html.Div(className="container", children=[
        html.Div(className="header", children=[
            # Note: We use root-relative paths for assets in modular apps
            html.Img(src='/assets/25h Logos.png', alt="RiskWatch Logo", className="app-logo"),
            html.P("Safety Observation Assistant", className="app-subtitle", style={'fontSize': '24px', 'fontWeight': '500'}),
            dcc.Link('Back to Home', href='/', className='nav-link')
        ]),
        html.Ul(id='flash-messages-container', className="flash-messages"),
        html.Div([
            html.Div(className="form-group", children=[
                html.Label("Floor:", htmlFor="floor-input"),
                dcc.Input(type="text", id="floor-input", placeholder="e.g., Ground Floor, B1, lvl 2...", required=True)
            ]),
            html.Div(className="form-group", children=[
                html.Label("Location:", htmlFor="location-input"),
                dcc.Input(type="text", id="location-input", placeholder="e.g., Main Lobby...", required=True)
            ]),
            html.Div(className="form-group", children=[
                html.Label("Observation Details:", htmlFor="observation-textarea"),
                dcc.Textarea(id="observation-textarea", placeholder="Describe what you observed...", required=True, rows=5)
            ]),
            html.Div(className="form-group", children=[
                html.Label("Attach Photo (Optional):"),
                dcc.Upload(
                    id='photo-upload', className="photo-upload-area",
                    children=html.A([
                        html.Img(src='/assets/upload-icon.png', className="photo-upload-img-icon"),
                        "Upload or Take Photo"
                    ], className="upload-photo-button-label"),
                    multiple=False
                ),
                html.Span(id="selected-file-name")
            ]),
            html.Div(className="button-group", children=[
                html.Button("Add Observation to Database", id="add-button", n_clicks=0, className="button-style"),
                dcc.Link('View Full Report', href='/report', className='button-style download-button', style={'textAlign': 'center'})
            ])
        ]),
        html.Div("© 2024 RiskWatch - Created by Arjun Menon", className="app-footer")
    ])

def build_report_page():
    """Builds the layout for the full observation report page."""
    return html.Div(className="report-container", children=[
        dcc.Download(id='download-excel'),
        html.Div(className="header", children=[
            html.Img(src='/assets/25h Logos.png', alt="RiskWatch Logo", className="app-logo"),
            html.H1("Full Safety Observation Report"),
            dcc.Link('Back to Home', href='/', className='nav-link')
        ]),
        html.Div(className="report-controls", children=[
            dcc.Input(id='search-input', type='text', placeholder='Search in descriptions, locations, floors...', debounce=True, className='search-bar'),
            html.Div(className="sort-dropdown-wrapper", children=[
                dcc.Dropdown(
                    id='sort-dropdown',
                    options=[
                        {'label': 'Sort by Newest First', 'value': 'date_newest'},
                        {'label': 'Sort by Oldest First', 'value': 'date_oldest'},
                        {'label': 'Sort by Highest Risk', 'value': 'risk_high'},
                    ], value='date_newest', clearable=False
                )
            ]),
            html.Button("Download Full Report as Excel", id='download-report-button', className="button-style download-button")
        ]),
        html.Div(id='report-content-container') # Content is loaded by a callback
    ])

# --- Helper Function for Excel Generation ---
def generate_excel_for_download(observations_data):
    # This function is unchanged and moved here for co-location with its usage.
    # ... (exact same code as in the original app.py)
    EXCEL_PHOTO_TARGET_WIDTH_PX = 150
    EXCEL_PHOTO_TARGET_HEIGHT_PX = 112
    EXCEL_ROW_HEIGHT_FOR_PHOTO_PT = 90.0
    EXCEL_PHOTO_COLUMN_WIDTH_UNITS = 22

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Safety Observation Report"

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    orange_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_fill = PatternFill(start_color="9C0006", end_color="9C0006", fill_type="solid")

    sheet.merge_cells('A1:C4')
    logo_path = os.path.join('assets', '25h Logos.png')
    if os.path.exists(logo_path):
        logo = OpenpyxlImage(logo_path)
        logo.height = 45; logo.width = 120
        sheet.add_image(logo, 'A1')

    title_font = Font(name='Calibri', size=16, bold=True, color="000080")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")

    sheet.merge_cells('D1:O4')
    title_cell = sheet['D1']
    title_cell.value = "SAFETY OBSERVATION REPORT"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    header_row = 5
    headers = ["ObsNo.", "Date of Observation", "Floor", "Location", "Description", "Impact", "Likelihood", "Severity", "Risk Rating", "Corrective Action Required", "Responsible Person", "Deadline", "Photo Evidence", "Closed Photo", "Status"]
    for col_idx, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        cell.value = header_text; cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    col_widths = {'A': 8, 'B': 20, 'C': 15, 'D': 35, 'E': 60, 'F': 45, 'G': 12, 'H': 12, 'I': 12, 'J': 60, 'K': 20, 'L': 15, 'M': EXCEL_PHOTO_COLUMN_WIDTH_UNITS, 'N': 20, 'O': 12}
    for col_letter, width in col_widths.items(): sheet.column_dimensions[col_letter].width = width
    for entry in observations_data:
        risk_rating = entry.get('risk_rating', 0)
        row_values = [entry.get('id'), entry.get('date_str'), entry.get('floor'), entry.get('location'), entry.get('description'), entry.get('impact'), entry.get('likelihood'), entry.get('severity'), risk_rating, entry.get('corrective_action'), entry.get('responsible_person', '').title(), entry.get('deadline'), None, "Attach closed photo", "Open"]
        sheet.append(row_values)
        new_row_num = sheet.max_row
        sheet.row_dimensions[new_row_num].height = EXCEL_ROW_HEIGHT_FOR_PHOTO_PT
        for col_idx_loop in range(1, len(headers) + 1):
            sheet.cell(row=new_row_num, column=col_idx_loop).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        risk_cell = sheet.cell(row=new_row_num, column=headers.index("Risk Rating") + 1)
        if 1 <= risk_rating <= 4: risk_cell.fill = green_fill
        elif 5 <= risk_rating <= 9: risk_cell.fill = yellow_fill
        elif 10 <= risk_rating <= 15: risk_cell.fill = orange_fill
        elif risk_rating >= 16: risk_cell.fill = red_fill
        if entry.get('photo_bytes'):
            try:
                img = OpenpyxlImage(io.BytesIO(entry['photo_bytes'])); img.width, img.height = EXCEL_PHOTO_TARGET_WIDTH_PX, EXCEL_PHOTO_TARGET_HEIGHT_PX
                sheet.add_image(img, get_column_letter(headers.index('Photo Evidence') + 1) + str(new_row_num))
            except Exception as e: print(f"Error embedding photo: {e}"); sheet.cell(row=new_row_num, column=(headers.index('Photo Evidence') + 1)).value = "Error"
        else: sheet.cell(row=new_row_num, column=(headers.index('Photo Evidence') + 1)).value = "No Photo"
    excel_stream = io.BytesIO(); workbook.save(excel_stream); excel_stream.seek(0)
    return excel_stream


# --- Callback Registration Function ---
def register_callbacks(app):
    """Registers all callbacks for the observation app."""

    @app.callback(
        Output('flash-messages-container', 'children'),
        Output('floor-input', 'value'),
        Output('location-input', 'value'),
        Output('observation-textarea', 'value'),
        Output('photo-upload', 'contents'),
        Output('selected-file-name', 'children', allow_duplicate=True),
        Input('add-button', 'n_clicks'),
        State('floor-input', 'value'),
        State('location-input', 'value'),
        State('observation-textarea', 'value'),
        State('photo-upload', 'contents'),
        prevent_initial_call=True
    )
    def add_observation(n_clicks, floor_input, location, observation, photo_contents):
        if not all([floor_input, location, observation]):
            return html.Li("Floor, Location, and Observation fields are required.", className="warning"), no_update, no_update, no_update, no_update, no_update
        photo_bytes = base64.b64decode(photo_contents.split(',')[1]) if photo_contents else None
        ai_analysis = get_ai_analysis(observation, floor_input, location)
        new_entry = {'date_str': datetime.datetime.now().strftime("%d-%b-%Y"), 'floor_from_user': floor_input, 'location_from_user': location, 'ai_analysis': ai_analysis, 'photo_bytes': photo_bytes}
        last_id = database.add_observation_to_db(new_entry)
        success_message = html.Li(f"Observation #{last_id} successfully saved.", className="success")
        return success_message, '', '', '', None, ''

    @app.callback(
        Output('report-content-container', 'children'),
        Input('url', 'pathname'),
        Input('search-input', 'value'),
        Input('sort-dropdown', 'value')
    )
    def update_report_view(pathname, search_term, sort_by):
        if pathname != '/report': raise PreventUpdate
        observations = database.get_observations_from_db(search_term, sort_by)
        if not observations: return html.P("No observations found.", style={'textAlign': 'center', 'padding': '50px'})
        cards = []
        for obs in observations:
            risk = obs.get('risk_rating', 0)
            risk_class = 'risk-low'
            if 5 <= risk <= 9: risk_class = 'risk-medium'
            elif 10 <= risk <= 15: risk_class = 'risk-high'
            elif risk >= 16: risk_class = 'risk-critical'
            card = html.Div(className="obs-card", children=[
                html.Div(className="card-main", children=[
                    html.H3(f"Obs #{obs['id']}: {obs['location']} ({obs['floor']})"),
                    html.P([html.B("Date: "), obs['date_str']]),
                    html.P([html.B("Impact: "), obs['impact']]),
                    html.P([html.B("Description: "), obs['description']]),
                    html.P([html.B("Corrective Action: "), obs['corrective_action']]),
                    html.P([html.B("Assigned To: "), f"{str(obs.get('responsible_person', 'N/A')).title()} | ", html.B("Deadline: "), f"{obs.get('deadline', 'N/A')}"])
                ]),
                html.Div(className="card-sidebar", children=[
                    html.Div(className="risk-box", children=[html.P("Risk Rating", className="risk-title"), html.P(risk, className=f"risk-value {risk_class}")]),
                    html.Img(src=f"data:image/png;base64,{obs['photo_b64']}" if obs['photo_b64'] else '/assets/placeholder.png', className="card-photo")
                ])
            ])
            cards.append(card)
        return cards

    @app.callback(Output('download-excel', 'data'), Input('download-report-button', 'n_clicks'), prevent_initial_call=True)
    def download_full_report(n_clicks):
        conn = database.get_db_connection()
        obs_for_excel = [dict(row) for row in conn.execute("SELECT * FROM observations ORDER BY id ASC").fetchall()]
        conn.close()
        if not obs_for_excel: raise PreventUpdate
        excel_stream = generate_excel_for_download(obs_for_excel)
        filename = f"Full_Safety_Report_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
        return dcc.send_bytes(excel_stream.read(), filename)

    @app.callback(Output('selected-file-name', 'children'), Input('photo-upload', 'filename'), prevent_initial_call=True)
    def update_filename_display(filename):
        return f"File selected: {filename}" if filename else ""